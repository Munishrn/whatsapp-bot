import openpyxl
import threading
from datetime import datetime
from config import EXCEL_FILE, STAFF_NUMBERS

excel_lock = threading.Lock()

COL_ID       = 0
COL_NAME     = 1
COL_DESC     = 2
COL_STATUS   = 3
COL_PHONE    = 4
COL_DATE     = 5
COL_DELIVERY = 6
COL_CREATED  = 7  # creation timestamp DD-MM-YYYY HH:MM


def get_user_role(phone):
    return "staff" if phone in STAFF_NUMBERS else "customer"


def generate_order_id(sheet):
    max_id = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).isdigit():
            max_id = max(max_id, int(row[0]))
    return str(max_id + 1)


def format_phone(number):
    number = str(number).strip()
    if not number.isdigit():
        return None
    if len(number) == 10:
        return "91" + number
    if len(number) == 12 and number.startswith("91"):
        return number
    return None


def is_ready(status):
    return status.strip().lower() == "ready to be picked"


def is_out_for_delivery(status):
    return status.strip().lower() == "out for delivery"


def is_cancelled(status):
    return status.strip().lower() == "cancelled"



def get_today_str():
    return datetime.now().strftime("%d-%m-%Y")


def parse_date(text):
    text = text.strip()
    formats = ["%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d"]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).strftime("%d-%m-%Y")
        except ValueError:
            continue
    return None


def parse_delivery_datetime(text):
    """Parse delivery date+time. Returns datetime object or None."""
    text = text.strip()
    formats = [
        "%d-%m-%Y %H:%M",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y %I:%M %p",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def format_delivery_str(dt):
    return dt.strftime("%d-%m-%Y %I:%M %p")


def get_customer_phone(order_id):
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            for row in sh.iter_rows(min_row=2, values_only=True):
                if str(row[COL_ID]) == str(order_id):
                    return str(row[COL_PHONE])
        except FileNotFoundError:
            print(f"Excel file not found: {EXCEL_FILE}")
    return None


def get_order_status(order_id, phone):
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            for row in sh.iter_rows(min_row=2, values_only=True):
                if str(row[COL_ID]) == str(order_id):
                    delivery = str(row[COL_DELIVERY]) if len(row) > COL_DELIVERY and row[COL_DELIVERY] else "N/A"
                    date_val = str(row[COL_DATE]) if len(row) > COL_DATE and row[COL_DATE] else "N/A"

                    if phone in STAFF_NUMBERS:
                        return (
                            f"Order ID: {row[COL_ID]}\n"
                            f"Customer: {row[COL_NAME]}\n"
                            f"Description: {row[COL_DESC]}\n"
                            f"Status: {row[COL_STATUS]}\n"
                            f"Phone: {row[COL_PHONE]}\n"
                            f"Created Date: {date_val}\n"
                            f"Expected Delivery: {delivery}"
                        )

                    if str(row[COL_PHONE]) != str(phone):
                        return "❌ Not authorized to view this order. It belongs to some other customer."

                    # ✅ Hide Expected Delivery for final statuses
                    hide_delivery = {"ready to be picked", "out for delivery", "cancelled"}
                    status_lower  = str(row[COL_STATUS]).strip().lower()
                    msg = (
                        f"Order ID: {row[COL_ID]}\n"
                        f"Customer: {row[COL_NAME]}\n"
                        f"Description: {row[COL_DESC]}\n"
                        f"Status: {row[COL_STATUS]}\n"
                        f"Created Date: {date_val}"
                    )
                    if status_lower not in hide_delivery and delivery != "N/A":
                        msg += f"\nExpected Delivery: {delivery}"
                    return msg
        except FileNotFoundError:
            return "❌ Order system not available. Please try again later."
    return "Order not found ❌"


def get_orders_by_date(date_str, phone, role):
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            matched = []
            for row in sh.iter_rows(min_row=2, values_only=True):
                if not row[COL_ID]:
                    continue
                row_date = str(row[COL_DATE]).strip() if len(row) > COL_DATE and row[COL_DATE] else ""
                if row_date != date_str:
                    continue
                delivery = str(row[COL_DELIVERY]) if len(row) > COL_DELIVERY and row[COL_DELIVERY] else "N/A"
                if role == "staff":
                    matched.append(
                        f"🆔 {row[COL_ID]} | 👤 {row[COL_NAME]} | 📦 {row[COL_DESC]}\n"
                        f"   Status: {row[COL_STATUS]} | 🕐 {delivery}"
                    )
                else:
                    if str(row[COL_PHONE]) == str(phone):
                        matched.append(
                            f"🆔 {row[COL_ID]} | 📦 {row[COL_DESC]}\n"
                            f"   Status: {row[COL_STATUS]} | 🕐 {delivery}"
                        )
            if not matched:
                return f"No orders found for {date_str} 📭"
            header = f"📅 Orders for {date_str} ({len(matched)} found):\n"
            header += "─" * 30 + "\n"
            return header + "\n\n".join(matched)
        except FileNotFoundError:
            return "❌ Order system not available. Please try again later."


def create_order(name, description, status, phone, delivery_str=""):
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sh = wb.active
            sh.append(["Order ID", "Customer", "Description", "Status", "Phone", "Date", "Delivery"])

        sh = wb.active
        if sh.cell(1, 6).value != "Date":
            sh.cell(1, 6).value = "Date"
        if sh.cell(1, 7).value != "Delivery":
            sh.cell(1, 7).value = "Delivery"

        order_id    = generate_order_id(sh)
        today       = get_today_str()
        created_at  = datetime.now().strftime("%d-%m-%Y %H:%M")  # ✅ Save creation timestamp
        if sh.cell(1, 8).value != "Created At":
            sh.cell(1, 8).value = "Created At"
        sh.append([order_id, name, description, status, phone, today, delivery_str, created_at])
        wb.save(EXCEL_FILE)
    return order_id


def update_order(order_id, status, delivery_str=None):
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            if sh.cell(1, 7).value != "Delivery":
                sh.cell(1, 7).value = "Delivery"
            for row in sh.iter_rows(min_row=2):
                if str(row[COL_ID].value) == str(order_id):
                    row[COL_STATUS].value = status
                    if delivery_str is not None:
                        row[COL_DELIVERY].value = delivery_str
                    wb.save(EXCEL_FILE)
                    return True
        except FileNotFoundError:
            print(f"Excel file not found: {EXCEL_FILE}")
    return False


def get_late_orders():
    """Returns orders past their delivery time that are not Ready or Cancelled."""
    now = datetime.now()
    skip_statuses = {"ready to be picked", "out for delivery", "cancelled"}
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            late = []
            for row in sh.iter_rows(min_row=2, values_only=True):
                if not row[COL_ID]:
                    continue
                status   = str(row[COL_STATUS]).strip().lower() if row[COL_STATUS] else ""
                delivery = str(row[COL_DELIVERY]).strip() if len(row) > COL_DELIVERY and row[COL_DELIVERY] else ""
                if status in skip_statuses or not delivery:
                    continue
                dt = parse_delivery_datetime(delivery)
                if dt and dt < now:
                    late.append({
                        "id":          row[COL_ID],
                        "customer":    row[COL_NAME],
                        "description": row[COL_DESC],
                        "status":      row[COL_STATUS],
                        "phone":       str(row[COL_PHONE]),
                        "delivery":    delivery,
                    })
            return late
        except FileNotFoundError:
            return []


def get_unique_clients():
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            seen = {}
            for row in sh.iter_rows(min_row=2, values_only=True):
                if row[COL_NAME] and row[COL_PHONE]:
                    name  = str(row[COL_NAME]).strip()
                    phone = str(row[COL_PHONE]).strip()
                    if name:
                        seen[name] = phone
            return [{"name": k, "phone": v} for k, v in seen.items()][:100]
        except FileNotFoundError:
            return []


def get_stale_orders(hours=6):
    skip_statuses = {"ready to be picked", "out for delivery", "cancelled"}
    cutoff = datetime.now()
    from datetime import timedelta
    threshold = cutoff - timedelta(hours=hours)

    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            stale = []
            for row in sh.iter_rows(min_row=2, values_only=True):
                if not row[COL_ID]:
                    continue
                status     = str(row[COL_STATUS]).strip().lower() if row[COL_STATUS] else ""
                created_at = str(row[COL_CREATED]).strip() if len(row) > COL_CREATED and row[COL_CREATED] else ""

                if status in skip_statuses:
                    continue

                # ✅ Only flag as stale if creation time is known and older than threshold
                if not created_at:
                    continue

                try:
                    created_dt = datetime.strptime(created_at, "%d-%m-%Y %H:%M")
                except ValueError:
                    continue

                if created_dt > threshold:
                    continue  # Not yet 6 hours old — skip

                stale.append({
                    "id":          row[COL_ID],
                    "customer":    row[COL_NAME],
                    "description": row[COL_DESC],
                    "status":      row[COL_STATUS],
                    "phone":       row[COL_PHONE],
                    "created_at":  created_at,
                })
            return stale
        except FileNotFoundError:
            return []
