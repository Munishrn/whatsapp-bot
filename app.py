import traceback
import shelve
import threading
import time

from flask import Flask, request

from logic import (
    get_user_role,
    get_order_status,
    get_orders_by_date,
    create_order,
    update_order,
    get_customer_phone,
    format_phone,
    is_ready,
    is_out_for_delivery,
    is_cancelled,
    get_today_str,
    parse_date,
    parse_delivery_datetime,
    format_delivery_str,
    get_unique_clients,
    get_stale_orders,
    get_late_orders,
)

from whatsapp import (
    send_staff_menu,
    send_customer_menu,
    send_text,
    send_back_button,
    send_status_list,
    send_date_options,
    send_view_options,
    send_create_order_type,
    send_existing_clients_list,
    send_update_options,
    send_delivery_date_picker,
    send_delivery_time_picker,
)

from config import VERIFY_TOKEN, STATUS_MAP, STAFF_NUMBERS

app = Flask(__name__)

STATE_FILE = "user_state"
TEMP_FILE  = "temp_data"

# Track orders already notified for late delivery to avoid repeat messages
notified_late = set()


# ── State helpers ─────────────────────────────────────────────────────────────

def get_state(phone):
    with shelve.open(STATE_FILE) as db:
        return db.get(phone)


def set_state(phone, state):
    with shelve.open(STATE_FILE) as db:
        db[phone] = state


def clear_state(phone):
    with shelve.open(STATE_FILE) as db:
        db.pop(phone, None)
    with shelve.open(TEMP_FILE) as db:
        db.pop(phone, None)


def get_temp(phone):
    with shelve.open(TEMP_FILE) as db:
        return db.get(phone, {})


def set_temp(phone, data):
    with shelve.open(TEMP_FILE) as db:
        db[phone] = data


def update_temp(phone, key, value):
    with shelve.open(TEMP_FILE) as db:
        current = db.get(phone, {})
        current[key] = value
        db[phone] = current


# ── Notification helpers ──────────────────────────────────────────────────────

def notify_ready(phone, order_id, description=""):
    msg = f"_🚚 Your order is Ready to be picked!_\n_Order ID: {order_id}_"
    if description:
        msg += f"\n_Product: {description}_"
    send_text(phone, msg)


def notify_cancelled(phone, order_id, description=""):
    msg = f"_❌ We're sorry to inform you that your order has been cancelled._\n_Order ID: {order_id}_"
    if description:
        msg += f"\n_Product: {description}_"
    msg += "\n_Please contact us for more details._"
    send_text(phone, msg)


def notify_status_update(phone, order_id, status, description="", delivery="", changed="status"):
    """Send auto notification to customer. changed= status | delivery | both"""
    status_emojis = {
        "design making":      "🎨",
        "plate making":       "🖼️",
        "offset printing":    "🖨️",
        "ready to be picked": "✅",
        "out for delivery":   "🚚",
        "cancelled":          "❌",
    }
    emoji = status_emojis.get(status.strip().lower(), "📋")
    msg = f"_{emoji} Order Update!_\n_Order ID: {order_id}_"
    if description:
        msg += f"\n_Product: {description}_"
    # ✅ Clearly mention what changed
    if changed == "status":
        msg += f"\n_✏️ Order status changed to: {status}_"
    elif changed == "delivery":
        msg += f"\n_Status: {status}_"
        if delivery:
            msg += f"\n_🕐 Expected delivery date/time changed to: {delivery}_"
    elif changed == "both":
        msg += f"\n_✏️ Order status changed to: {status}_"
        if delivery:
            msg += f"\n_🕐 Expected delivery date/time changed to: {delivery}_"
    send_text(phone, msg)


# ── Late delivery background checker ─────────────────────────────────────────

def late_delivery_checker():
    """
    Runs in background thread every 30 minutes.
    Checks for orders past their delivery time and notifies customers.
    """
    while True:
        try:
            late_orders = get_late_orders()
            for order in late_orders:
                order_id = str(order["id"])
                if order_id not in notified_late:
                    send_text(
                        order["phone"],
                        f"_⏰ We apologize for the delay on your order._\n"
                        f"_Order ID: {order_id}_\n"
                        f"_Product: {order['description']}_\n"
                        f"_Expected Delivery: {order['delivery']}_\n"
                        f"_Current Status: {order['status']}_\n\n"
                        f"_Our team is working on it. We will update you shortly. 🙏_"
                    )
                    notified_late.add(order_id)
                    print(f"[Late Delivery] Notified customer for Order {order_id}")
        except Exception as e:
            print(f"[Late Delivery Checker Error] {e}")

        time.sleep(30 * 60)  # Check every 30 minutes


# Start background thread
checker_thread = threading.Thread(target=late_delivery_checker, daemon=True)
checker_thread.start()
print("[Late Delivery Checker] Started — checking every 30 minutes.")


# ── Webhook ───────────────────────────────────────────────────────────────────
@app.route("/", methods=["GET"])
def health():
    return "Bot is running!", 200
    
@app.route("/webhook", methods=["GET"])
def verify():
    if request.args.get("hub.verify_token") == VERIFY_TOKEN:
        return request.args.get("hub.challenge")
    return "Verification failed", 403


@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.json

    try:
        entry = data.get("entry", [])
        if not entry:
            return "ok"
        changes = entry[0].get("changes", [])
        if not changes:
            return "ok"
        value = changes[0].get("value", {})
        messages = value.get("messages", [])
        if not messages:
            return "ok"

        message = messages[0]
        phone   = message["from"]
        role    = get_user_role(phone)
        state   = get_state(phone)

        # ── INTERACTIVE ───────────────────────────────────────────────────
        if "interactive" in message:
            interactive = message["interactive"]

            if "button_reply" in interactive:
                button_id = interactive["button_reply"]["id"]
            elif "list_reply" in interactive:
                button_id = interactive["list_reply"]["id"]
            else:
                return "ok"

            # Back to menu
            if button_id == "back_to_menu":
                clear_state(phone)
                if role == "staff":
                    send_staff_menu(phone)
                else:
                    send_customer_menu(phone)

            # Create order
            elif button_id == "create_order":
                send_create_order_type(phone)

            elif button_id == "client_new":
                set_state(phone, "create_name")
                send_back_button(phone, "Enter Customer Name:")

            elif button_id == "client_existing":
                clients = get_unique_clients()
                if not clients:
                    send_text(phone, "❌ No existing clients found. Please create a new client.")
                    send_create_order_type(phone)
                else:
                    set_state(phone, "select_existing_client")
                    send_existing_clients_list(phone, clients)

            # Update order
            elif button_id == "update_order":
                set_state(phone, "update_order_id")
                send_back_button(phone, "Enter Order ID to update:")

            # What to update — status, delivery, or both
            elif button_id == "update_status_only" and state == "update_what":
                update_temp(phone, "update_type", "status")
                set_state(phone, "update_status")
                send_status_list(phone, f"Select new status for Order {get_temp(phone).get('order_id')}:")

            elif button_id == "update_delivery_only" and state == "update_what":
                update_temp(phone, "update_type", "delivery")
                set_state(phone, "update_delivery_date")
                send_delivery_date_picker(phone)

            elif button_id == "update_both" and state == "update_what":
                update_temp(phone, "update_type", "both")
                set_state(phone, "update_status")
                send_status_list(phone, f"Select new status for Order {get_temp(phone).get('order_id')}:")

            # View orders
            elif button_id == "view_orders":
                send_view_options(phone)

            elif button_id == "view_order":
                set_state(phone, "view_order")
                send_back_button(phone, "Enter Order ID to view:")

            elif button_id == "view_by_date":
                set_state(phone, "view_by_date")
                send_date_options(phone)

            elif button_id.startswith("filter_date_") and state == "view_by_date":
                date_val = button_id.replace("filter_date_", "")
                if date_val == "custom":
                    set_state(phone, "enter_specific_date")
                    send_back_button(phone, "Enter date as DD-MM-YYYY\nExample: 28-04-2026")
                else:
                    result = get_orders_by_date(date_val, phone, role)
                    send_back_button(phone, result)
                    clear_state(phone)

            # Customer track order
            elif button_id == "check_status":
                set_state(phone, "check_status")
                send_text(phone, "Enter your Order ID:")

            # ── Delivery date selected from picker ──────────────────────
            elif button_id.startswith("del_date_"):
                date_val = button_id.replace("del_date_", "")
                if date_val == "custom":
                    # Ask staff to type custom date
                    if state and state.startswith("update"):
                        set_state(phone, "update_delivery_custom_date")
                    else:
                        set_state(phone, "create_delivery_custom_date")
                    send_back_button(phone, "Enter custom date as DD-MM-YYYY\nExample: 05-05-2026\nNote: Must be today or a future date")
                else:
                    # Date selected — now pick time
                    update_temp(phone, "delivery_date", date_val)
                    if state and state.startswith("update"):
                        set_state(phone, "update_delivery_time")
                    else:
                        set_state(phone, "create_delivery_time")
                    send_delivery_time_picker(phone, date_val)

            # ── Delivery time selected from picker ──────────────────────
            elif button_id.startswith("del_time_"):
                time_val = button_id.replace("del_time_", "")
                d = get_temp(phone)

                if time_val == "custom":
                    # Custom time — ask to type HH:MM manually
                    d = get_temp(phone)
                    delivery_date = d.get("delivery_date", "")
                    if state and state.startswith("update"):
                        set_state(phone, "update_delivery_custom_time")
                    else:
                        set_state(phone, "create_delivery_custom_time")
                    send_back_button(phone, f"Enter custom time for {delivery_date} as HH:MM\nExample: 14:30 for 2:30 PM\n16:45 for 4:45 PM")
                else:
                    delivery_date = d.get("delivery_date", "")
                    dt_str = f"{delivery_date} {time_val}"
                    dt = parse_delivery_datetime(dt_str)

                    # ✅ Validate — reject if selected time is in the past
                    from datetime import datetime as dt_class
                    if dt and dt <= dt_class.now():
                        send_delivery_time_picker(phone, delivery_date)
                        send_text(phone, "⚠️ That time has already passed. Please select a future time slot.")
                        return "ok"

                    delivery_str = format_delivery_str(dt) if dt else dt_str

                    if state and state.startswith("update"):
                        order_id    = d.get("order_id")
                        update_type = d.get("update_type", "delivery")
                        new_status  = d.get("new_status")
                        _apply_update(phone, order_id, update_type, new_status, delivery_str)
                    else:
                        # Create order
                        order_id = create_order(d["name"], d["description"], d["status"], d["phone"], delivery_str)
                        send_text(
                            d["phone"],
                            f"_📦 Order Created!_\n"
                            f"_Order ID: {order_id}_\n"
                            f"_Product: {d['description']}_\n"
                            f"_Status: {d['status']}_\n"
                            f"_Expected Delivery: {delivery_str}_"
                        )
                        send_text(phone, f"✅ Order Created\nOrder ID: {order_id}\nExpected Delivery: {delivery_str}")
                        clear_state(phone)

            # Existing client selected
            elif button_id.startswith("existing_client_") and state == "select_existing_client":
                parts = button_id.split("_", 3)
                client_phone = parts[3] if len(parts) > 3 else ""
                client_name  = interactive.get("list_reply", {}).get("title", "")
                set_temp(phone, {"name": client_name, "phone": client_phone})
                set_state(phone, "create_description")
                send_back_button(phone, f"✅ Client: {client_name}\nEnter Description/Product:")

            # Status selected during CREATE
            elif button_id.startswith("status_") and state == "create_status":
                status = STATUS_MAP.get(button_id)
                if not status:
                    send_text(phone, "❌ Invalid status selected.")
                    return "ok"
                update_temp(phone, "status", status)
                # ✅ Skip delivery picker for final statuses
                skip_delivery = {"ready to be picked", "out for delivery", "cancelled"}
                if status.strip().lower() in skip_delivery:
                    d = get_temp(phone)
                    order_id = create_order(d["name"], d["description"], status, d["phone"], "")
                    send_text(
                        d["phone"],
                        f"_📦 Order Created!_\n"
                        f"_Order ID: {order_id}_\n"
                        f"_Product: {d['description']}_\n"
                        f"_Status: {status}_"
                    )
                    send_text(phone, f"✅ Order Created\nOrder ID: {order_id}\nStatus: {status}")
                    if is_cancelled(status):
                        cp = get_customer_phone(order_id)
                        if cp:
                            notify_cancelled(cp, order_id, d["description"])
                    clear_state(phone)
                else:
                    set_state(phone, "create_delivery_date")
                    send_delivery_date_picker(phone)

            # Status selected during UPDATE
            elif button_id.startswith("status_") and state == "update_status":
                status = STATUS_MAP.get(button_id)
                if not status:
                    send_text(phone, "❌ Invalid status selected.")
                    return "ok"

                d        = get_temp(phone)
                order_id = d.get("order_id")
                update_type = d.get("update_type", "status")

                if update_type == "both":
                    # ✅ Skip delivery picker for final statuses
                    skip_delivery = {"ready to be picked", "out for delivery", "cancelled"}
                    if status.strip().lower() in skip_delivery:
                        update_temp(phone, "new_status", status)
                        order_id = get_temp(phone).get("order_id")
                        success = update_order(order_id, status)
                        if success:
                            send_text(phone, f"✅ Order {order_id} updated to: {status}")
                            cp   = get_customer_phone(order_id)
                            desc = get_order_desc(order_id)
                            if cp:
                                if is_cancelled(status):
                                    notify_cancelled(cp, order_id, desc)
                                elif is_ready(status):
                                    notify_ready(cp, order_id, desc)
                                else:
                                    notify_status_update(cp, order_id, status, desc, "", changed="status")
                        else:
                            send_text(phone, f"❌ Order {order_id} not found.")
                        clear_state(phone)
                    else:
                        # Save status, then show date picker
                        update_temp(phone, "new_status", status)
                        set_state(phone, "update_delivery_date")
                        send_delivery_date_picker(phone)
                else:
                    # Status only — notify customer on every change
                    success = update_order(order_id, status)
                    if success:
                        send_text(phone, f"✅ Order {order_id} updated to: {status}")
                        cp   = get_customer_phone(order_id)
                        desc = get_order_desc(order_id)
                        if cp:
                            if is_ready(status):
                                notify_ready(cp, order_id, desc)
                            elif is_cancelled(status):
                                notify_cancelled(cp, order_id, desc)
                            else:
                                delivery = get_order_delivery(order_id)
                                notify_status_update(cp, order_id, status, desc, delivery, changed="status")
                    else:
                        send_text(phone, f"❌ Order {order_id} not found.")
                    clear_state(phone)

        # ── TEXT ──────────────────────────────────────────────────────────
        elif "text" in message:
            text = message["text"]["body"].strip()

            hinglish_greetings = [
                "hi", "hello", "menu", "start", "hii", "hiii", "hey",
                "namaste", "namaskar", "namsate", "sat sri akal",
                "sat shri akal", "ssa", "salaam", "adaab", "hy", "helo",
                "order", "help", "madad", "kya hai", "kya h", "bhai", "ji",
                "hello ji", "hi ji", "hey ji",
            ]

            if text.lower().strip() in hinglish_greetings:
                clear_state(phone)
                if role == "staff":
                    stale = get_stale_orders(hours=6)
                    if stale:
                        lines = [f"⚠️ Stale Orders Alert — {len(stale)} order(s) not updated in 6+ hours:\n"]
                        for o in stale:
                            lines.append(
                                f"🆔 Order {o['id']} | 👤 {o['customer']} | 📦 {o['description']}\n"
                                f"   Status: {o['status']} | 🕐 Created: {o.get('created_at', 'N/A')}"
                            )
                        send_text(phone, "\n".join(lines))
                    send_staff_menu(phone)
                else:
                    send_customer_menu(phone)

            elif state == "enter_specific_date":
                parsed = parse_date(text)
                if not parsed:
                    # ✅ Use send_text not send_back_button — Back button would clear state
                    send_text(phone, "❌ Invalid date format. Please enter as DD-MM-YYYY\nExample: 28-04-2026")
                    return "ok"
                result = get_orders_by_date(parsed, phone, role)
                send_back_button(phone, result)
                clear_state(phone)

            elif state == "create_name":
                set_temp(phone, {"name": text})
                set_state(phone, "create_description")
                send_back_button(phone, "Enter Description/Product:")

            elif state == "create_description":
                update_temp(phone, "description", text)
                d = get_temp(phone)
                if d.get("phone"):
                    set_state(phone, "create_status")
                    send_status_list(phone, "Select Order Status:")
                else:
                    set_state(phone, "create_phone")
                    send_back_button(phone, "Enter Customer Phone Number:")

            elif state == "create_phone":
                formatted = format_phone(text)
                if not formatted:
                    send_back_button(phone, "❌ Invalid number. Enter 10-digit or 12-digit with country code:")
                    return "ok"
                update_temp(phone, "phone", formatted)
                set_state(phone, "create_status")
                send_status_list(phone, "Select Order Status:")

            elif state == "create_delivery_custom_date":
                # Staff entered a custom date manually
                from logic import parse_date as pd
                parsed = pd(text)
                if not parsed:
                    send_back_button(phone, "❌ Invalid date. Enter as DD-MM-YYYY\nExample: 30-04-2026")
                    return "ok"
                update_temp(phone, "delivery_date", parsed)
                set_state(phone, "create_delivery_time")
                send_delivery_time_picker(phone, parsed)

            elif state == "create_delivery_custom_time":
                from datetime import datetime as dt_class
                d = get_temp(phone)
                dt_str = f"{d['delivery_date']} {text.strip()}"
                dt = parse_delivery_datetime(dt_str)
                if not dt:
                    send_back_button(phone, "❌ Invalid time. Enter as HH:MM\nExample: 14:00")
                    return "ok"
                # ✅ Validate past time
                if dt <= dt_class.now():
                    send_back_button(phone, "⚠️ That time has already passed.\nEnter a future time as HH:MM\nExample: 14:00")
                    return "ok"
                delivery_str = format_delivery_str(dt)
                order_id = create_order(d["name"], d["description"], d["status"], d["phone"], delivery_str)
                send_text(
                    d["phone"],
                    f"_📦 Order Created!_\n"
                    f"_Order ID: {order_id}_\n"
                    f"_Product: {d['description']}_\n"
                    f"_Status: {d['status']}_\n"
                    f"_Expected Delivery: {delivery_str}_"
                )
                send_text(phone, f"✅ Order Created\nOrder ID: {order_id}\nExpected Delivery: {delivery_str}")
                clear_state(phone)

            elif state == "update_order_id":
                check = get_order_status(text, phone)
                if "not found" in check.lower():
                    send_back_button(phone, f"❌ Order {text} not found.\n\nPlease enter a valid Order ID:")
                else:
                    # ✅ Check if order is in a final locked status
                    current = get_current_status(text)
                    locked_statuses = {"ready to be picked", "out for delivery", "cancelled"}
                    if current.strip().lower() in locked_statuses:
                        send_back_button(phone, f"🔒 Order {text} is in '{current}' status and cannot be updated further.")
                    else:
                        set_temp(phone, {"order_id": text})
                        set_state(phone, "update_what")
                        send_update_options(phone)

            elif state == "update_delivery_custom_date":
                from logic import parse_date as pd
                parsed = pd(text)
                if not parsed:
                    send_back_button(phone, "❌ Invalid date. Enter as DD-MM-YYYY\nExample: 30-04-2026")
                    return "ok"
                update_temp(phone, "delivery_date", parsed)
                set_state(phone, "update_delivery_time")
                send_delivery_time_picker(phone, parsed)

            elif state == "update_delivery_custom_time":
                from datetime import datetime as dt_class
                d = get_temp(phone)
                dt_str = f"{d['delivery_date']} {text.strip()}"
                dt = parse_delivery_datetime(dt_str)
                if not dt:
                    send_back_button(phone, "❌ Invalid time. Enter as HH:MM\nExample: 14:00")
                    return "ok"
                # ✅ Validate past time
                if dt <= dt_class.now():
                    send_back_button(phone, "⚠️ That time has already passed.\nEnter a future time as HH:MM\nExample: 14:00")
                    return "ok"
                delivery_str = format_delivery_str(dt)
                order_id = d.get("order_id")
                update_type = d.get("update_type", "delivery")
                _apply_update(phone, order_id, update_type, d.get("new_status"), delivery_str)

            elif state == "view_order":
                result = get_order_status(text, phone)
                if "not found" in result.lower():
                    send_back_button(phone, result + "\n\nPlease enter a valid Order ID:")
                else:
                    send_back_button(phone, result)
                    clear_state(phone)

            elif state == "check_status":
                raw_result = get_order_status(text, phone)
                if "❌" not in raw_result and "not found" not in raw_result.lower():
                    send_text(phone, raw_result)
                    clear_state(phone)
                else:
                    send_back_button(phone, raw_result + "\n\nPlease enter a valid Order ID:")

            elif role == "customer" and text.strip().isdigit():
                raw_result = get_order_status(text.strip(), phone)
                if "❌" not in raw_result and "not found" not in raw_result.lower():
                    send_text(phone, raw_result)
                    clear_state(phone)
                else:
                    set_state(phone, "check_status")
                    send_back_button(phone, raw_result + "\n\nPlease enter a valid Order ID:")

            else:
                if role == "staff":
                    send_staff_menu(phone)
                else:
                    send_customer_menu(phone)

    except Exception:
        print(f"[Webhook Error]\n{traceback.format_exc()}")

    return "ok"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _apply_update(phone, order_id, update_type, new_status, delivery_str):
    """Apply order update and send notifications."""
    if update_type == "both":
        success = update_order(order_id, new_status, delivery_str)
        if success:
            send_text(phone, f"✅ Order {order_id} updated\nStatus: {new_status}\nExpected Delivery: {delivery_str}")
            cp   = get_customer_phone(order_id)
            desc = get_order_desc(order_id)
            if cp:
                if is_ready(new_status):
                    notify_ready(cp, order_id, desc)
                elif is_cancelled(new_status):
                    notify_cancelled(cp, order_id, desc)
                else:
                    # Use delivery_str directly — already updated in Excel
                    notify_status_update(cp, order_id, new_status, desc, delivery_str, changed="both")
        else:
            send_text(phone, f"❌ Order {order_id} not found.")
    else:
        # Delivery only — notify customer with updated delivery time
        current_status = get_current_status(order_id)
        success = update_order(order_id, current_status, delivery_str)
        if success:
            send_text(phone, f"✅ Delivery time updated for Order {order_id}\nNew Expected Delivery: {delivery_str}")
            cp   = get_customer_phone(order_id)
            desc = get_order_desc(order_id)
            if cp:
                notify_status_update(cp, order_id, current_status, desc, delivery_str, changed="delivery")
        else:
            send_text(phone, f"❌ Order {order_id} not found.")
    clear_state(phone)


def get_order_desc(order_id):
    """Get description for an order — used in notifications."""
    import openpyxl
    from config import EXCEL_FILE
    from logic import excel_lock, COL_ID, COL_DESC
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            for row in sh.iter_rows(min_row=2, values_only=True):
                if str(row[COL_ID]) == str(order_id):
                    return str(row[COL_DESC]) if row[COL_DESC] else ""
        except Exception:
            pass
    return ""


def get_order_delivery(order_id):
    """Get expected delivery time for an order."""
    import openpyxl
    from config import EXCEL_FILE
    from logic import excel_lock, COL_ID, COL_DELIVERY
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            for row in sh.iter_rows(min_row=2, values_only=True):
                if str(row[COL_ID]) == str(order_id):
                    return str(row[COL_DELIVERY]) if len(row) > COL_DELIVERY and row[COL_DELIVERY] else ""
        except Exception:
            pass
    return ""


def get_current_status(order_id):
    """Get current status of an order — used when updating delivery only."""
    import openpyxl
    from config import EXCEL_FILE
    from logic import excel_lock, COL_ID, COL_STATUS
    with excel_lock:
        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            sh = wb.active
            for row in sh.iter_rows(min_row=2, values_only=True):
                if str(row[COL_ID]) == str(order_id):
                    return str(row[COL_STATUS]) if row[COL_STATUS] else "Order Received"
        except Exception:
            pass
    return "Order Received"


if __name__ == "__main__":
    import os
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)), debug=False)
